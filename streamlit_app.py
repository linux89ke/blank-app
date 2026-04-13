"""
Decathlon Product Lookup
Category matching modes:
  - Keyword (TF-IDF over deca_cat.xlsx)   — always available, instant
  - AI (TF-IDF shortlist → Groq rerank)   — toggle on, needs Groq API key
"""

import os, io, re, json, asyncio
import numpy as np
import streamlit as st
import pandas as pd
import requests
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# ── Optional Groq import ──────────────────────────────────────────────────────
try:
    from groq import AsyncGroq, Groq as SyncGroq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Decathlon Product Lookup", page_icon="🏅", layout="wide")
st.markdown("""
<style>
h1 { color: #0082C3; }
.tag {
    display:inline-block; background:#0082C3; color:white;
    border-radius:4px; padding:2px 8px; font-size:12px; margin:2px;
}
.ai-badge {
    display:inline-block; background:linear-gradient(90deg,#f55036,#ff8c00);
    color:white; border-radius:12px; padding:2px 10px;
    font-size:11px; font-weight:700; margin-left:6px;
}
.kw-badge {
    display:inline-block; background:#0082C3; color:white;
    border-radius:12px; padding:2px 10px;
    font-size:11px; font-weight:700; margin-left:6px;
}
</style>
""", unsafe_allow_html=True)

st.title("🏅 Decathlon Product Lookup")
st.markdown("Search by model number or product name — view details, images, and **download a filled upload template**.")

# ── Constants ─────────────────────────────────────────────────────────────────
IMAGE_COLS    = ["OG_image"] + [f"picture_{i}" for i in range(1, 11)]
TEMPLATE_PATH = "product-creation-template.xlsx"
DECA_CAT_PATH = "deca_cat.xlsx"
MASTER_PATH   = "Decathlon_Working_File_Split.csv"

MASTER_TO_TEMPLATE = {
    "product_name":   "Name",
    "designed_for":   "Description",
    "sku_num_sku_r3": "SellerSKU",
    "model_code":     "ParentSKU",
    "brand_name":     "Brand",
    "bar_code":       "GTIN_Barcode",
    "color":          "color",
    "model_label":    "model",
    "keywords":       "note",
    "weight":         "product_weight",
    "OG_image":       "MainImage",
    "picture_1":      "Image2",
    "picture_2":      "Image3",
    "picture_3":      "Image4",
    "picture_4":      "Image5",
    "picture_5":      "Image6",
    "picture_6":      "Image7",
    "picture_7":      "Image8",
}

# Fields combined into the product query string for category matching
CATEGORY_MATCH_FIELDS = [
    "department_label",  # e.g. "SOCCER / FUTSAL"
    "nature_label",      # e.g. "SPORTS ITEM"
    "product_name",      # e.g. "Kids' Long-Sleeved Thermal Base Layer"
    "channable_gender",  # e.g. "BOYS'|GIRLS'"
    "family",            # e.g. "UNDER TS JR"
    "type",              # e.g. "JUNIOR FOOTBALL"
    "brand_name",        # e.g. "KIPSTA"
    "keywords",          # e.g. "Thermal Tops|Kids|Clothing|..."
    "model_label",       # e.g. "KEEPDRY DDY LS BLACK"
]

# ── Groq prompt ───────────────────────────────────────────────────────────────
GROQ_SYSTEM = """You are a product categorization expert for a sports retailer.
Given a product description and candidate category paths, pick the {top_n} best matches.
Consider brand, product type, gender, sport, and age group.

Respond with JSON only:
{{
  "categories": [
    {{"category": "<full path>", "score": 0.95}},
    ...
  ]
}}

Rules:
- Return exactly {top_n} categories ordered by confidence descending
- Only pick from the provided candidate list - never invent categories
- Scores are floats 0.0-1.0
- JSON only, nothing else"""

# =============================================================================
# DATA LOADING
# =============================================================================

@st.cache_data(show_spinner=False)
def load_reference_data(file_bytes: bytes):
    wb_bytes = io.BytesIO(file_bytes)
    df_cat = pd.read_excel(wb_bytes, sheet_name="category", dtype=str)
    df_cat.columns = [c.strip() for c in df_cat.columns]
    df_cat = df_cat[df_cat["export_category"].notna() & (df_cat["export_category"].str.strip() != "")]
    df_cat["export_category"]       = df_cat["export_category"].str.strip()
    df_cat["category_name_lower"]   = df_cat["category_name"].str.lower().str.strip()
    df_cat["Category Path lower"]   = df_cat["Category Path"].str.lower().fillna("")

    wb_bytes.seek(0)
    df_brands = pd.read_excel(wb_bytes, sheet_name="brands", dtype=str, header=0)
    df_brands.columns = ["brand_entry"]
    df_brands = df_brands[df_brands["brand_entry"].notna()].copy()
    df_brands["brand_entry"]        = df_brands["brand_entry"].str.strip()
    df_brands["brand_name_lower"]   = (
        df_brands["brand_entry"].str.split(" - ", n=1).str[-1].str.lower().str.strip()
    )
    return df_cat, df_brands


@st.cache_data(show_spinner=False)
def load_master(file_bytes: bytes, is_csv: bool) -> pd.DataFrame:
    if is_csv:
        try:
            return pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            return pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="latin-1")
    return pd.read_excel(io.BytesIO(file_bytes), dtype=str)


# =============================================================================
# TF-IDF INDEX  (built once from category sheet, cached)
# =============================================================================

def _path_to_doc(path: str) -> str:
    """Repeat leaf segments to upweight specificity."""
    parts = path.split(" / ")
    return " ".join(parts) + " " + " ".join(parts[-3:]) * 2


@st.cache_resource(show_spinner=False)
def build_tfidf_index(ref_bytes: bytes):
    """Build TF-IDF index over leaf category paths from deca_cat.xlsx."""
    df_cat, _ = load_reference_data(ref_bytes)
    all_paths  = df_cat["Category Path"].dropna().astype(str).tolist()
    path_set   = set(all_paths)
    leaves     = [p for p in all_paths
                  if not any(other.startswith(p + " / ") for other in path_set)]
    docs       = [_path_to_doc(p) for p in leaves]
    vectorizer = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True)
    matrix     = vectorizer.fit_transform(docs)
    path_to_export = dict(zip(df_cat["Category Path"], df_cat["export_category"]))
    return leaves, vectorizer, matrix, path_to_export


def tfidf_shortlist(queries: list, leaves, vectorizer, matrix, k: int = 30) -> list:
    """Return top-k leaf paths per query in one matrix op."""
    qmat = vectorizer.transform(queries)
    sims = cosine_similarity(qmat, matrix)
    out  = []
    for row in sims:
        top_idx = np.argsort(row)[::-1][:k]
        out.append([leaves[i] for i in top_idx if row[i] > 0])
    return out


# =============================================================================
# KEYWORD MATCHING  (fast, no API)
# =============================================================================

def _build_query_string(row: pd.Series) -> str:
    parts = []
    for f in CATEGORY_MATCH_FIELDS:
        val = row.get(f, "")
        if pd.notna(val) and str(val).strip() not in ("", "-", "nan"):
            parts.append(str(val).strip().lower())
    return " ".join(parts)


def keyword_match_category(row: pd.Series, df_cat: pd.DataFrame) -> tuple:
    """
    Score every category by token overlap + depth tiebreaker.
    Returns (primary_export_category, additional_export_category).
    """
    query = _build_query_string(row)
    if not query:
        return "", ""
    q_tokens = set(re.findall(r"[a-z]+", query))

    def score(cat_row):
        p_tokens = set(re.findall(r"[a-z]+", cat_row["Category Path lower"]))
        overlap  = len(q_tokens & p_tokens)
        depth    = cat_row["Category Path lower"].count("/")
        bonus    = 2 if cat_row["category_name_lower"] in query else 0
        return overlap + bonus + depth * 0.1

    scores  = df_cat.apply(score, axis=1)
    top_idx = scores.nlargest(2).index
    results = df_cat.loc[top_idx, "export_category"].tolist()
    return (results[0] if results else ""), (results[1] if len(results) > 1 else "")


# =============================================================================
# AI MATCHING  (TF-IDF shortlist -> Groq rerank)
# =============================================================================

async def _async_rerank(idx, query, candidates, client, model, top_n, sem):
    async with sem:
        cand_list = "\n".join(f"- {c}" for c in candidates)
        try:
            resp = await client.chat.completions.create(
                model=model,
                temperature=0.1,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": GROQ_SYSTEM.format(top_n=top_n)},
                    {"role": "user",   "content": f"Product: {query}\n\nCandidates:\n{cand_list}"},
                ],
            )
            data = json.loads(resp.choices[0].message.content.strip()).get("categories", [])
            return idx, data
        except Exception as e:
            return idx, [{"category": f"ERROR: {e}", "score": 0.0}]


async def _parallel_rerank(queries, candidates_list, api_key, model, top_n, concurrency):
    client = AsyncGroq(api_key=api_key)
    sem    = asyncio.Semaphore(concurrency)
    tasks  = [
        _async_rerank(i, q, c, client, model, top_n, sem)
        for i, (q, c) in enumerate(zip(queries, candidates_list))
    ]
    raw = await asyncio.gather(*tasks)
    return [r for _, r in sorted(raw, key=lambda x: x[0])]


def groq_rerank_batch(queries, candidates_list, api_key, model, top_n, concurrency):
    return asyncio.run(
        _parallel_rerank(queries, candidates_list, api_key, model, top_n, concurrency)
    )


def ai_match_categories(
    rows_df, leaves, vectorizer, matrix, path_to_export,
    api_key, model, shortlist_k=30, concurrency=10,
) -> list:
    """Run TF-IDF shortlist then Groq rerank for every row. Returns [(primary, additional)]."""
    queries         = [_build_query_string(row) for _, row in rows_df.iterrows()]
    candidates_list = tfidf_shortlist(queries, leaves, vectorizer, matrix, shortlist_k)
    all_preds       = groq_rerank_batch(
        queries, candidates_list, api_key, model, top_n=2, concurrency=concurrency
    )

    def _resolve(cat_path: str) -> str:
        if cat_path in path_to_export:
            return path_to_export[cat_path]
        for p, ex in path_to_export.items():
            if p.endswith(cat_path) or cat_path.endswith(p):
                return ex
        return cat_path

    results = []
    for preds in all_preds:
        primary   = _resolve(preds[0]["category"]) if len(preds) > 0 else ""
        secondary = _resolve(preds[1]["category"]) if len(preds) > 1 else ""
        results.append((primary, secondary))
    return results


# =============================================================================
# BRAND MATCHING
# =============================================================================

def match_brand(raw: str, df_brands: pd.DataFrame) -> str:
    if not raw or pd.isna(raw):
        return ""
    needle = str(raw).strip().lower()
    exact  = df_brands[df_brands["brand_name_lower"] == needle]
    if not exact.empty:
        return exact.iloc[0]["brand_entry"]
    partial = df_brands[df_brands["brand_name_lower"].str.contains(needle, regex=False)]
    if not partial.empty:
        return partial.iloc[0]["brand_entry"]
    for _, brow in df_brands.iterrows():
        if brow["brand_name_lower"] in needle:
            return brow["brand_entry"]
    return str(raw).strip()


# =============================================================================
# TEMPLATE BUILDER
# =============================================================================

def build_template(results_df, df_cat, df_brands, use_ai, ai_categories) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Upload Template"]

    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            header_map[val] = col_idx

    hfont      = ws.cell(row=1, column=1).font
    data_font  = Font(name=hfont.name or "Calibri", size=hfont.size or 11)
    data_align = Alignment(vertical="center")

    for i, (_, src_row) in enumerate(results_df.iterrows()):
        row_idx  = i + 2
        row_data = {}

        for master_col, tmpl_col in MASTER_TO_TEMPLATE.items():
            val = src_row.get(master_col, "")
            if pd.notna(val) and str(val).strip() not in ("", "nan"):
                row_data[tmpl_col] = str(val).strip()

        raw_brand = src_row.get("brand_name", "")
        if pd.notna(raw_brand) and str(raw_brand).strip():
            row_data["Brand"] = match_brand(str(raw_brand), df_brands)

        if use_ai and ai_categories and i < len(ai_categories):
            primary, secondary = ai_categories[i]
        else:
            primary, secondary = keyword_match_category(src_row, df_cat)

        if primary:
            row_data["PrimaryCategory"]    = primary
        if secondary:
            row_data["AdditionalCategory"] = secondary

        for tmpl_col, value in row_data.items():
            if tmpl_col in header_map:
                cell           = ws.cell(row=row_idx, column=header_map[tmpl_col])
                cell.value     = value
                cell.font      = data_font
                cell.alignment = data_align

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# SIDEBAR
# =============================================================================

with st.sidebar:
    st.header("📂 Reference Data")
    uploaded_ref = st.file_uploader("deca_cat.xlsx (categories & brands)", type=["xlsx"])

    st.header("📂 Master Data")
    uploaded_master = st.file_uploader("Working file (.xlsx or .csv)", type=["xlsx", "csv"])

    st.markdown("---")

    # Category matching mode toggle
    st.header("🧠 Category Matching")
    use_ai_matching = st.toggle(
        "AI matching (Groq)",
        value=False,
        help=(
            "OFF = fast keyword/TF-IDF matching against category paths. "
            "ON = TF-IDF shortlist + Groq LLM rerank (more accurate, requires free API key)."
        ),
    )

    if use_ai_matching:
        if not GROQ_AVAILABLE:
            st.error("Install groq: `pip install groq`")
            use_ai_matching = False
        else:
            st.markdown('<span class="ai-badge">AI MODE ON</span>', unsafe_allow_html=True)
            groq_api_key = st.text_input(
                "Groq API key", type="password",
                value=os.environ.get("GROQ_API_KEY", ""),
                placeholder="gsk_...",
            )
            st.caption("Free key at [console.groq.com](https://console.groq.com)")
            groq_model = st.selectbox(
                "Model",
                ["llama-3.1-8b-instant", "llama-3.3-70b-versatile", "mixtral-8x7b-32768"],
                index=0,
                help="8b = fastest & free. 70b = most accurate.",
            )
            shortlist_k = st.slider("Shortlist size (candidates/product)", 10, 50, 30)
            concurrency = st.slider("Parallel Groq requests", 1, 30, 10)
    else:
        st.markdown('<span class="kw-badge">KEYWORD MODE</span>', unsafe_allow_html=True)
        st.caption("Instant TF-IDF keyword overlap matching. No API key needed.")
        groq_api_key = ""
        groq_model   = "llama-3.1-8b-instant"
        shortlist_k  = 30
        concurrency  = 10

    st.markdown("---")
    st.header("🔎 Search Fields")
    search_fields = st.multiselect(
        "Match terms against",
        ["model_code", "model_label", "product_name", "sku_num_sku_r3", "Jumia SKU", "bar_code"],
        default=["model_code", "model_label", "product_name"],
    )
    st.markdown("---")
    show_images = st.checkbox("Show product images", value=True)
    max_images  = st.slider("Max images per product", 1, 11, 5)


# =============================================================================
# LOAD DATA
# =============================================================================

# Reference data
if uploaded_ref:
    ref_bytes = uploaded_ref.read()
    st.sidebar.success("✅ Custom deca_cat.xlsx loaded")
else:
    try:
        ref_bytes = open(DECA_CAT_PATH, "rb").read()
        st.sidebar.info("📋 Using bundled deca_cat.xlsx")
    except FileNotFoundError:
        ref_bytes = None
        st.sidebar.warning("⚠️ deca_cat.xlsx not found - upload in sidebar")

if ref_bytes:
    df_cat, df_brands = load_reference_data(ref_bytes)
    st.sidebar.success(f"✅ {len(df_cat):,} categories · {len(df_brands)} brands")
    leaves, vectorizer, tfidf_matrix, path_to_export = build_tfidf_index(ref_bytes)
else:
    df_cat = df_brands = leaves = vectorizer = tfidf_matrix = path_to_export = None

# Master data
if uploaded_master:
    master_bytes = uploaded_master.read()
    is_csv       = uploaded_master.name.endswith(".csv")
    df_master    = load_master(master_bytes, is_csv)
    st.sidebar.success(f"✅ {len(df_master):,} product rows loaded")
else:
    loaded = False
    for path, csv in [(MASTER_PATH, True), (MASTER_PATH.replace(".csv", ".xlsx"), False)]:
        try:
            master_bytes = open(path, "rb").read()
            df_master    = load_master(master_bytes, csv)
            st.sidebar.info(f"📋 Bundled master · {len(df_master):,} rows")
            loaded = True
            break
        except FileNotFoundError:
            continue
    if not loaded:
        st.error("No master file found. Upload one in the sidebar.")
        st.stop()

img_cols_present = [c for c in IMAGE_COLS if c in df_master.columns]
data_cols        = [c for c in df_master.columns if c not in img_cols_present]


# =============================================================================
# SEARCH
# =============================================================================

def search(q: str) -> pd.DataFrame:
    mask = pd.Series(False, index=df_master.index)
    for field in search_fields:
        if field in df_master.columns:
            mask |= df_master[field].fillna("").str.lower().str.contains(q.lower(), regex=False)
    return df_master[mask].copy()


# =============================================================================
# INPUT TABS
# =============================================================================

tab1, tab2 = st.tabs(["📤 Upload a List", "⌨️ Manual Entry"])
queries = []

with tab1:
    uploaded_list = st.file_uploader(
        "Upload file with model numbers / product names",
        type=["xlsx", "csv", "txt"],
        help="One value per row. For Excel/CSV, values must be in column A.",
    )
    if uploaded_list:
        ext = uploaded_list.name.rsplit(".", 1)[-1].lower()
        if ext == "txt":
            queries = [l.strip() for l in uploaded_list.read().decode().splitlines() if l.strip()]
        elif ext == "csv":
            q_df    = pd.read_csv(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        else:
            q_df    = pd.read_excel(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        st.success(f"Loaded **{len(queries)}** search terms")

with tab2:
    manual = st.text_area(
        "Enter one model number or product name per line",
        height=160,
        placeholder="8641696\nKEEPDRY DDY LS BLACK\n4271703",
    )
    if manual.strip():
        queries = [q.strip() for q in manual.strip().splitlines() if q.strip()]


# =============================================================================
# RESULTS
# =============================================================================

if queries:
    st.markdown("---")
    all_result_frames = []
    no_match          = []

    for q in queries:
        res = search(q)
        if res.empty:
            no_match.append(q)
        else:
            res.insert(0, "Search Term", q)
            all_result_frames.append((q, res))

    if no_match:
        st.warning(f"No matches found for: **{', '.join(no_match)}**")

    if all_result_frames:
        total_rows = sum(len(r) for _, r in all_result_frames)
        st.success(f"**{total_rows} rows** matched across **{len(all_result_frames)}** query(ies)")

        combined = pd.concat([r for _, r in all_result_frames], ignore_index=True)

        # ── AI category matching (batch, one Groq call per product in parallel) ──
        ai_categories = None

        if df_cat is not None and use_ai_matching:
            if not groq_api_key:
                st.warning("Enter your Groq API key in the sidebar to use AI matching.")
                use_ai_matching = False
            else:
                n = len(combined)
                est = max(2, n // concurrency + 2)
                with st.spinner(
                    f"🤖 AI matching {n} products with Groq "
                    f"({concurrency} parallel calls, ~{est}s)..."
                ):
                    try:
                        ai_categories = ai_match_categories(
                            combined, leaves, vectorizer, tfidf_matrix, path_to_export,
                            groq_api_key, groq_model, shortlist_k, concurrency,
                        )
                        st.success(f"✅ AI matched {n} products")
                    except Exception as e:
                        st.error(f"Groq error: {e}")
                        ai_categories   = None
                        use_ai_matching = False

        # ── Preview ────────────────────────────────────────────────────────────
        if df_cat is not None:
            mode_label = "🤖 AI" if (use_ai_matching and ai_categories) else "🔑 Keyword"
            with st.expander(f"{mode_label} — Category & Brand Matching Preview", expanded=False):
                preview_rows = []
                for i, (_, prow) in enumerate(combined.iterrows()):
                    if use_ai_matching and ai_categories:
                        prim, addl = ai_categories[i]
                    else:
                        prim, addl = keyword_match_category(prow, df_cat)
                    preview_rows.append({
                        "SKU":                prow.get("sku_num_sku_r3", ""),
                        "Product":            str(prow.get("product_name", ""))[:60],
                        "Brand (raw)":        prow.get("brand_name", ""),
                        "Brand (matched)":    match_brand(str(prow.get("brand_name", "")), df_brands),
                        "PrimaryCategory":    prim,
                        "AdditionalCategory": addl,
                        "Method":             "AI" if (use_ai_matching and ai_categories) else "Keyword",
                    })
                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

        # ── Download buttons ───────────────────────────────────────────────────
        col_dl1, col_dl2 = st.columns(2)

        with col_dl1:
            raw_out = io.BytesIO()
            with pd.ExcelWriter(raw_out, engine="openpyxl") as writer:
                combined.to_excel(writer, index=False, sheet_name="Results")
            st.download_button(
                "⬇️ Download Raw Results (.xlsx)",
                data=raw_out.getvalue(),
                file_name="decathlon_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col_dl2:
            if df_cat is None:
                st.warning("Upload deca_cat.xlsx to enable the filled template download.")
            else:
                try:
                    tpl_bytes  = build_template(
                        combined, df_cat, df_brands,
                        use_ai=bool(use_ai_matching and ai_categories),
                        ai_categories=ai_categories,
                    )
                    mode_icon  = "🤖" if (use_ai_matching and ai_categories) else "🔑"
                    st.download_button(
                        f"{mode_icon} Download Filled Upload Template (.xlsx)",
                        data=tpl_bytes,
                        file_name="decathlon_upload_template_filled.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                except FileNotFoundError:
                    st.warning(
                        "Template file not found. "
                        "Place `product-creation-template.xlsx` in the app folder."
                    )

        st.markdown("---")

        # ── Per-query result cards ─────────────────────────────────────────────
        for q, res in all_result_frames:
            with st.expander(f"🔍 **{q}**  —  {len(res)} row(s)", expanded=True):
                show_cols = ["Search Term"] + [c for c in data_cols if c in res.columns]
                st.dataframe(
                    res[show_cols],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "keywords":     st.column_config.TextColumn("keywords",     width="large"),
                        "product_name": st.column_config.TextColumn("product_name", width="large"),
                        "designed_for": st.column_config.TextColumn("designed_for", width="large"),
                    },
                )

                cats = set()
                for _, row in res.iterrows():
                    for col in ["department_label", "nature_label", "family", "type"]:
                        val = row.get(col, "")
                        if pd.notna(val) and str(val).strip():
                            cats.add(str(val).strip())
                if cats:
                    tags = " ".join(f'<span class="tag">{c}</span>' for c in sorted(cats))
                    st.markdown(f"**Categories & Types:** {tags}", unsafe_allow_html=True)

                if show_images and img_cols_present:
                    first_row = res.iloc[0]
                    img_urls  = [
                        str(first_row[c]) for c in img_cols_present
                        if pd.notna(first_row.get(c))
                        and str(first_row.get(c, "")).startswith("http")
                    ][:max_images]
                    if img_urls:
                        st.markdown("**🖼 Product Images**")
                        cols = st.columns(len(img_urls))
                        for i, url in enumerate(img_urls):
                            try:
                                resp = requests.get(url, timeout=6)
                                img  = Image.open(io.BytesIO(resp.content))
                                cols[i].image(
                                    img,
                                    caption="Main" if i == 0 else f"View {i}",
                                    use_container_width=True,
                                )
                            except Exception:
                                cols[i].markdown(f"[🔗 Image {i+1}]({url})")
else:
    st.info("👆 Upload a list or type search terms above to get started.")

st.markdown("---")
st.caption("Decathlon Product Lookup · Powered by your Decathlon working file")
